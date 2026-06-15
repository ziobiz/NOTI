# -*- coding: utf-8 -*-
import json
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd


def parse_csv(path):
    return pd.read_csv(path, encoding="utf-8-sig", dtype=str).fillna("")


def norm_tx(s):
    s = str(s or "").strip()
    if re.match(r"^\d+\.0$", s):
        s = s.split(".")[0]
    return s


def load_icopay_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if c is None else str(c).strip()) for c in row])
        sheets[name] = rows
    wb.close()
    return sheets


def rows_to_df(rows):
    if not rows:
        return pd.DataFrame()
    header = rows[0]
    # skip empty header rows
    start = 0
    for i, r in enumerate(rows[:5]):
        if any(c for c in r):
            header = r
            start = i
            break
    data = rows[start + 1 :]
    # trim to header width
    w = len(header)
    data = [r[:w] + [""] * max(0, w - len(r)) for r in data if any(str(x).strip() for x in r)]
    df = pd.DataFrame(data, columns=header)
    return df


def find_col(df, *candidates):
    cols = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = cand.lower()
        if k in cols:
            return cols[k]
        for ck, orig in cols.items():
            if k in ck or ck in k:
                return orig
    return None


def status_is_success_icopay(val):
    s = str(val or "").strip().lower()
    return s in ("success", "paid", "결제", "성공", "complete", "0", "1", "승인", "approved")


def status_is_success_pg(val):
    return str(val or "").strip().lower() == "success"


def status_is_paid_noti(val):
    return str(val or "").strip() == "결제"


def main():
    icopay_path = Path(sys.argv[1])
    noti_path = Path(sys.argv[2])
    pg_path = Path(sys.argv[3])

    sheets = load_icopay_xlsx(icopay_path)
    # use first sheet with most rows, or sheet named like payment
    best_name = max(sheets.keys(), key=lambda n: len(sheets[n]))
    for n in sheets:
        if "결제" in n or "payment" in n.lower() or "내역" in n:
            best_name = n
            break
    icopay_df = rows_to_df(sheets[best_name])
    noti_df = parse_csv(noti_path)
    pg_df = parse_csv(pg_path)

    print("ICOPAY sheet:", best_name)
    print("ICOPAY columns:", list(icopay_df.columns)[:30])
    print("ICOPAY rows:", len(icopay_df))

    tx_col = find_col(
        icopay_df,
        "TransactionId",
        "transactionid",
        "transaction_id",
        "거래번호",
        "트랜잭션",
        "transno",
        "trans no",
    )
    st_col = find_col(
        icopay_df,
        "Status",
        "status",
        "상태",
        "결제상태",
        "paymentstatus",
        "처리상태",
    )
    date_col = find_col(icopay_df, "거래일", "결제일", "일자", "date", "거래일시", "paymentdate")

    if not tx_col:
        # guess: column with mostly numeric ids
        for c in icopay_df.columns:
            sample = icopay_df[c].head(50).astype(str)
            if sample.str.match(r"^\d{6,}$").mean() > 0.5:
                tx_col = c
                break

    icopay = {}
    for _, r in icopay_df.iterrows():
        tx = norm_tx(r.get(tx_col, "") if tx_col else "")
        if not tx:
            continue
        st = r.get(st_col, "") if st_col else ""
        icopay[tx] = {
            "status": st,
            "date": r.get(date_col, "") if date_col else "",
            "success": status_is_success_icopay(st),
        }

    noti = {}
    for _, r in noti_df.iterrows():
        tx = norm_tx(r.get("TransactionId", ""))
        if not tx:
            continue
        noti[tx] = {"status": r.get("상태", ""), "success": status_is_paid_noti(r.get("상태", ""))}

    pg = {}
    for _, r in pg_df.iterrows():
        tx = norm_tx(r.get("TransactionId", ""))
        if not tx:
            continue
        pg[tx] = {"status": r.get("Status", ""), "success": status_is_success_pg(r.get("Status", ""))}

    icopay_tx = set(icopay.keys())
    noti_tx = set(noti.keys())
    pg_tx = set(pg.keys())

    both3 = icopay_tx & noti_tx & pg_tx
    only_icopay = icopay_tx - noti_tx - pg_tx
    only_icopay_noti = (icopay_tx & noti_tx) - pg_tx
    only_icopay_pg = (icopay_tx & pg_tx) - noti_tx
    only_pg_noti = (pg_tx & noti_tx) - icopay_tx
    only_pg = pg_tx - icopay_tx - noti_tx
    only_noti = noti_tx - icopay_tx - pg_tx

    def count_success(d, tx_set):
        return sum(1 for t in tx_set if d.get(t, {}).get("success"))

    icopay_st_breakdown = {}
    for v in icopay.values():
        k = v["status"] or "(blank)"
        icopay_st_breakdown[k] = icopay_st_breakdown.get(k, 0) + 1

    mismatch_icopay_pg = []
    for tx in icopay_tx & pg_tx:
        i_ok = icopay[tx]["success"]
        p_ok = pg[tx]["success"]
        if i_ok != p_ok:
            mismatch_icopay_pg.append(
                {"tx": tx, "icopay": icopay[tx]["status"], "pg": pg[tx]["status"]}
            )

    mismatch_icopay_noti = []
    for tx in icopay_tx & noti_tx:
        i_ok = icopay[tx]["success"]
        n_ok = noti[tx]["success"]
        if i_ok != n_ok:
            mismatch_icopay_noti.append(
                {"tx": tx, "icopay": icopay[tx]["status"], "noti": noti[tx]["status"]}
            )

    out = {
        "icopay_sheet": best_name,
        "icopay_tx_col": tx_col,
        "icopay_status_col": st_col,
        "icopay_total_tx": len(icopay_tx),
        "noti_total_tx": len(noti_tx),
        "pg_total_tx": len(pg_tx),
        "icopay_success": count_success(icopay, icopay_tx),
        "noti_success_결제": count_success(noti, noti_tx),
        "pg_success": count_success(pg, pg_tx),
        "in_all_three": len(both3),
        "only_icopay": len(only_icopay),
        "only_icopay_and_noti": len(only_icopay_noti),
        "only_icopay_and_pg": len(only_icopay_pg),
        "only_pg_and_noti_not_icopay": len(only_pg_noti),
        "only_pg": len(only_pg),
        "only_noti": len(only_noti),
        "icopay_status_breakdown": dict(sorted(icopay_st_breakdown.items(), key=lambda x: -x[1])[:20]),
        "only_icopay_pg_success_sample": [
            {"tx": tx, "icopay": icopay[tx]["status"], "pg": pg[tx]["status"]}
            for tx in sorted(only_icopay_pg)[:20]
            if pg.get(tx, {}).get("success")
        ],
        "only_pg_not_icopay_success_sample": [
            {"tx": tx, "pg": pg[tx]["status"], "noti": noti.get(tx, {}).get("status", "")}
            for tx in sorted(only_pg_noti)[:20]
            if pg.get(tx, {}).get("success")
        ],
        "only_noti_not_icopay_success_sample": [
            {"tx": tx, "noti": noti[tx]["status"]}
            for tx in sorted(only_noti)[:20]
            if noti.get(tx, {}).get("success")
        ],
        "only_icopay_success_sample": [
            {"tx": tx, "status": icopay[tx]["status"]}
            for tx in sorted(only_icopay)[:20]
            if icopay.get(tx, {}).get("success")
        ],
        "mismatch_icopay_pg_count": len(mismatch_icopay_pg),
        "mismatch_icopay_noti_count": len(mismatch_icopay_noti),
        "gap_pg_success_minus_icopay": count_success(pg, pg_tx) - count_success(icopay, icopay_tx),
        "gap_icopay_success_minus_noti": count_success(icopay, icopay_tx) - count_success(noti, noti_tx),
    }
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
