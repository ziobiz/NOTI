# -*- coding: utf-8 -*-
"""Compare ChillPay paymenttransaction xlsx vs noti/pg CSV exports."""
import json
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd


def norm_tx(s):
    s = str(s or "").strip()
    if re.match(r"^\d+\.0$", s):
        s = s.split(".")[0]
    return s


def parse_csv(path):
    return pd.read_csv(path, encoding="utf-8-sig", dtype=str).fillna("")


def pg_ymd(s):
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", str(s or "").strip())
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else str(s)[:10]


def find_col(df, *candidates):
    cols = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = cand.lower()
        if k in cols:
            return cols[k]
        for ck, orig in cols.items():
            if k in ck or ck in k:
                return orig
    return None


def load_chillpay_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    best_name = wb.sheetnames[0]
    best_rows = 0
    for name in wb.sheetnames:
        ws = wb[name]
        n = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        if n > best_rows:
            best_rows = n
            best_name = name
    ws = wb[best_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([("" if c is None else str(c).strip()) for c in row])
    wb.close()
    if not rows:
        return best_name, pd.DataFrame(), []
    header_idx = 0
    for i, r in enumerate(rows[:20]):
        line = " ".join(r).lower()
        if "trans. id" in line or "trans id" in line or ("transaction" in line and "date" in line):
            header_idx = i
            break
    header = rows[header_idx]
    # dedupe column names (empty Merchant sub-column etc.)
    seen = {}
    clean_header = []
    for h in header:
        key = h.strip() or "_blank"
        seen[key] = seen.get(key, 0) + 1
        clean_header.append(key if seen[key] == 1 else f"{key}_{seen[key]}")
    data = [
        r[: len(header)] + [""] * max(0, len(header) - len(r))
        for r in rows[header_idx + 1 :]
        if any(x.strip() for x in r) and not str(r[0]).strip().lower().startswith("total")
    ]
    df = pd.DataFrame(data, columns=clean_header)
    return best_name, df, list(df.columns)


def status_success_chill(val):
    s = str(val or "").strip().lower()
    return s == "success"


def status_success_pg(val):
    return str(val or "").strip().lower() == "success"


def status_paid_noti(val):
    return str(val or "").strip() == "결제"


def build_map_chillpay(df):
    tx_col = find_col(
        df,
        "Trans. ID",
        "TransactionId",
        "Transaction ID",
        "transactionid",
        "transaction_id",
        "TransNo",
        "Trans No",
        "transno",
    )
    st_col = find_col(df, "Status", "status", "PaymentStatus", "paymentstatus", "State")
    date_col = find_col(
        df,
        "TransactionDate",
        "PaymentDate",
        "Date",
        "CreatedDate",
        "transaction date",
        "payment date",
    )
    if not tx_col:
        for c in df.columns:
            ser = df[c].head(30).astype(str)
            if ser.str.match(r"^\d{6,}$").mean() > 0.4:
                tx_col = c
                break
    out = {}
    for _, r in df.iterrows():
        tx = norm_tx(r.get(tx_col, "") if tx_col else "")
        if not tx:
            continue
        st = r.get(st_col, "") if st_col else ""
        out[tx] = {
            "status": st,
            "date": str(r.get(date_col, ""))[:10] if date_col else "",
            "success": status_success_chill(st),
        }
    return out, tx_col, st_col, date_col


def main():
    xlsx_path = Path(sys.argv[1])
    noti_path = Path(sys.argv[2])
    pg_path = Path(sys.argv[3])

    sheet, cdf, cols = load_chillpay_xlsx(xlsx_path)
    chill, tx_col, st_col, date_col = build_map_chillpay(cdf)

    noti_df = parse_csv(noti_path)
    pg_df = parse_csv(pg_path)
    noti = {
        norm_tx(r.TransactionId): {"status": r.get("상태", ""), "date": str(r.get("수신일", ""))[:10], "success": status_paid_noti(r.get("상태", ""))}
        for _, r in noti_df.iterrows()
        if norm_tx(r.TransactionId)
    }
    pg = {
        norm_tx(r.TransactionId): {"status": r.get("Status", ""), "date": pg_ymd(r.get("거래일시", "")), "success": status_success_pg(r.get("Status", ""))}
        for _, r in pg_df.iterrows()
        if norm_tx(r.TransactionId)
    }

    C, N, P = set(chill), set(noti), set(pg)

    def succ(d, s):
        return sum(1 for t in s if d.get(t, {}).get("success"))

    chill_st = Counter(v["status"] for v in chill.values())
    # refine success if status col ambiguous: also count common success labels in breakdown
    if succ(chill, C) == 0 and chill_st:
        # try PaymentStatus numeric
        for _, r in cdf.iterrows():
            tx = norm_tx(r.get(tx_col, "") if tx_col else "")
            if not tx or tx not in chill:
                continue
            for c in cdf.columns:
                v = str(r.get(c, "")).strip()
                if v in ("1", "Success", "success", "0"):
                    if c.lower().find("status") >= 0 or c.lower().find("payment") >= 0:
                        chill[tx]["success"] = v.lower() in ("success", "1", "0") or v == "0"

    only_c = C - N - P
    only_c_pg = (C & P) - N
    only_c_noti = (C & N) - P
    all3 = C & N & P
    only_pg = P - C
    only_noti = N - C

    mm_cp = [t for t in C & P if chill[t]["success"] != pg[t]["success"]]
    mm_cn = [t for t in C & N if chill[t]["success"] != noti[t]["success"]]
    mm_pn = [t for t in P & N if pg[t]["success"] != noti[t]["success"]]

    out = {
        "chillpay_xlsx": {
            "sheet": sheet,
            "rows": len(cdf),
            "unique_tx": len(C),
            "tx_col": tx_col,
            "status_col": st_col,
            "date_col": date_col,
            "columns": cols[:25],
            "status_breakdown": dict(chill_st.most_common(15)),
            "success_count": succ(chill, C),
        },
        "noti_csv": {"unique_tx": len(N), "success_결제": succ(noti, N), "status": dict(Counter(v["status"] for v in noti.values()))},
        "pg_csv": {"unique_tx": len(P), "success": succ(pg, P), "status": dict(Counter(v["status"] for v in pg.values()))},
        "overlap": {
            "all_three": len(all3),
            "chill_and_pg_not_noti": len(only_c_pg),
            "chill_and_noti_not_pg": len(only_c_noti),
            "only_chillpay": len(only_c),
            "only_pg_vs_chill": len(only_pg),
            "only_noti_vs_chill": len(only_noti),
            "pg_and_noti_not_chill": len((P & N) - C),
        },
        "gaps": {
            "chill_minus_pg_success": succ(chill, C) - succ(pg, P),
            "chill_minus_noti_success": succ(chill, C) - succ(noti, N),
            "pg_minus_noti_success": succ(pg, P) - succ(noti, N),
        },
        "mismatch_in_overlap": {
            "chill_vs_pg": len(mm_cp),
            "chill_vs_noti": len(mm_cn),
            "pg_vs_noti": len(mm_pn),
        },
        "only_chill_success_sample": [
            {"tx": t, **chill[t]}
            for t in sorted(only_c)
            if chill[t]["success"]
        ][:12],
        "only_pg_success_not_chill": [
            {"tx": t, **pg[t]}
            for t in sorted(only_pg)
            if pg[t]["success"]
        ][:12],
        "only_noti_success_not_chill": [
            {"tx": t, **noti[t]}
            for t in sorted(only_noti)
            if noti[t]["success"]
        ][:12],
        "chill_pg_not_noti_success": [
            {"tx": t, "chill": chill[t]["status"], "pg": pg[t]["status"]}
            for t in sorted(only_c_pg)
            if chill[t]["success"] or pg[t]["success"]
        ][:15],
        "sample_mismatch_chill_pg": [{"tx": t, "chill": chill[t], "pg": pg[t]} for t in mm_cp[:10]],
    }
    out_path = Path(__file__).parent / "compare-chillpay-xlsx-result.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(str(out_path))


if __name__ == "__main__":
    main()
