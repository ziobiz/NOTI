# -*- coding: utf-8 -*-
"""Compare ICOPAY dev xlsx vs 5/26 noti CSV vs pg CSV vs optional internal xlsx."""
import json
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd


def norm_tx(s):
    s = str(s or "").strip()
    if re.match(r"^\d+\.0$", s):
        s = s.split(".")[0]
    return s


def parse_csv(path):
    return pd.read_csv(path, encoding="utf-8-sig", dtype=str).fillna("")


def pg_ymd(s):
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", str(s or "").strip())
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else str(s)[:10]


def load_icopay_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[("" if c is None else c) for c in r] for r in ws.iter_rows(values_only=True)]
    wb.close()
    header_idx = 0
    for i, r in enumerate(rows[:20]):
        line = " ".join(str(x) for x in r)
        if "승인" in line or "거래일" in line or "주문" in line:
            header_idx = i
            break
    header = [str(x).strip() for x in rows[header_idx]]
    data = [
        r[: len(header)] + [""] * max(0, len(header) - len(r))
        for r in rows[header_idx + 1 :]
        if any(str(x).strip() for x in r)
    ]
    return pd.DataFrame(data, columns=header), header


def find_col(df, *cands):
    for c in df.columns:
        cs = str(c).strip()
        for cand in cands:
            if cs == cand or cand.lower() in cs.lower():
                return c
    return None


def load_tx_map(df, tx_col, date_col=None, status_col=None, amt_col=None, date_fn=None):
    out = {}
    if tx_col is None:
        return out
    for _, row in df.iterrows():
        tx = norm_tx(row.get(tx_col, ""))
        if not tx or not re.match(r"^\d+$", tx):
            continue
        rec = {}
        if date_col:
            d = str(row.get(date_col, "")).strip()
            rec["date"] = date_fn(d) if date_fn else d[:10]
        if status_col:
            rec["status"] = str(row.get(status_col, "")).strip()
        if amt_col:
            rec["amt"] = str(row.get(amt_col, "")).strip()
        out[tx] = rec
    return out


def main():
    dl = Path(r"c:\Users\ziobi\Downloads")
    icopay_path = dl / "결제내역_모든리스트_20260527 (1).xlsx"
    noti_path = dl / "transactions-2026-05-26 (1).csv"
    pg_path = dl / "pg-transactions-2026-05-26 (1).csv"
    # 전산(ICOPAY 운영) — 5/26 일자 단건 export (112건, 피지노티 CSV 건수와 동일)
    internal_path = dl / "결제내역_20260526.xlsx"
    if not internal_path.exists():
        internal_path = dl / "결제내역_20260526 (1).xlsx"

    icopay_df, _ = load_icopay_xlsx(icopay_path)
    noti_df = parse_csv(noti_path)
    pg_df = parse_csv(pg_path)

    ic_tx = find_col(icopay_df, "승인번호", "TransactionId")
    ic_date = find_col(icopay_df, "거래일")
    ic_status = find_col(icopay_df, "상태")
    ic_amt = find_col(icopay_df, "결제금액")

    icopay = load_tx_map(icopay_df, ic_tx, ic_date, ic_status, ic_amt)
    noti = load_tx_map(
        noti_df,
        find_col(noti_df, "TransactionId"),
        find_col(noti_df, "거래일", "수신일"),
        find_col(noti_df, "상태"),
        find_col(noti_df, "Amount"),
    )
    pg = load_tx_map(
        pg_df,
        find_col(pg_df, "TransactionId"),
        find_col(pg_df, "거래일시"),
        find_col(pg_df, "Status"),
        find_col(pg_df, "Amount"),
        date_fn=pg_ymd,
    )

    internal = {}
    if internal_path and internal_path.exists():
        int_df, _ = load_icopay_xlsx(internal_path)
        internal = load_tx_map(
            int_df,
            find_col(int_df, "승인번호", "TransactionId"),
            find_col(int_df, "거래일"),
            find_col(int_df, "상태"),
            find_col(int_df, "결제금액"),
        )

    # PG CSV는 5/26 피지 동기화분. ICOPAY·노티는 파일 전체(대부분 5/26).
    pg26 = {k: v for k, v in pg.items() if v.get("date") == "2026-05-26" or not v.get("date")}

    sets = {
        "icopay_dev": set(icopay),
        "noti_피지": set(noti),
        "chillpay_pg": set(pg26),
    }
    if internal:
        sets["icopay_전산"] = set(internal)

    def pairwise(all_sets):
        res = {}
        keys = list(all_sets)
        for i, a in enumerate(keys):
            for b in keys[i + 1 :]:
                res[f"only_{a}"] = res.get(f"only_{a}", {})
                res[f"both_{a}_{b}"] = len(all_sets[a] & all_sets[b])
                res[f"only_{a}_not_{b}"] = len(all_sets[a] - all_sets[b])
        return res

    keys = list(sets)
    pw = {}
    for i, a in enumerate(keys):
        for b in keys[i + 1 :]:
            pw[f"{a}|{b}"] = {
                "both": len(sets[a] & sets[b]),
                f"only_{a}": len(sets[a] - sets[b]),
                f"only_{b}": len(sets[b] - sets[a]),
                f"sample_only_{a}": sorted(sets[a] - sets[b])[:20],
                f"sample_only_{b}": sorted(sets[b] - sets[a])[:20],
            }

    all_keys = list(sets.keys())
    triple = sets[all_keys[0]]
    for k in all_keys[1:]:
        triple &= sets[k]

    out = {
        "files": {
            "icopay_dev": str(icopay_path),
            "noti": str(noti_path),
            "pg": str(pg_path),
            "internal_xlsx": str(internal_path) if internal_path else None,
        },
        "counts_total": {k: len(v) for k, v in {"icopay": icopay, "noti": noti, "pg": pg, "internal": internal}.items()},
        "counts_in_compare": {k: len(v) for k, v in sets.items()},
        "status_breakdown": {
            "icopay_dev": dict(Counter(v.get("status", "") for v in icopay.values())),
            "noti": dict(Counter(v.get("status", "") for v in noti.values())),
            "pg": dict(Counter(v.get("status", "") for v in pg26.values())),
        },
        "pairwise": pw,
        "in_all_sources": len(triple),
        "icopay_dev_only_not_noti_nor_pg": sorted(sets["icopay_dev"] - sets["noti_피지"] - sets["chillpay_pg"]),
        "pg_only_not_icopay_dev": sorted(sets["chillpay_pg"] - sets["icopay_dev"]),
        "noti_only_not_icopay_dev": sorted(sets["noti_피지"] - sets["icopay_dev"]),
        "noti_and_pg_not_icopay_dev": sorted(sets["noti_피지"] & sets["chillpay_pg"] - sets["icopay_dev"]),
    }
    if internal:
        out["status_breakdown"]["icopay_전산"] = dict(Counter(v.get("status", "") for v in internal.values()))
        out["dev_vs_전산"] = {
            "both": len(sets["icopay_dev"] & sets["icopay_전산"]),
            "dev_only": sorted(sets["icopay_dev"] - sets["icopay_전산"]),
            "전산_only": sorted(sets["icopay_전산"] - sets["icopay_dev"]),
        }
        out["전산_vs_noti"] = {
            "both": len(sets["icopay_전산"] & sets["noti_피지"]),
            "전산_only": sorted(sets["icopay_전산"] - sets["noti_피지"])[:30],
            "noti_only": sorted(sets["noti_피지"] - sets["icopay_전산"])[:30],
        }
        triple_all = sets["icopay_dev"] & sets["icopay_전산"] & sets["noti_피지"] & sets["chillpay_pg"]
        out["in_all_four"] = len(triple_all)

    out_path = Path(__file__).resolve().parent / "compare-four-20260526-result.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(out, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
